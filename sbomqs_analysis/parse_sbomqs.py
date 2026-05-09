#!/usr/bin/env python3
# Reads the sbomqs output saved by run_sbomqs.sh and writes to an existing Excel sheet.

import re
import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

BASE_PATH = "/rust_projects"


def read_column_map(ws):
    """
    Reads the two header rows from the Excel sheet and returns a map like:
      { "syft": { "sbomqs score": 2, "bsi-v2.1": 3, "required": 4, ... },
        "trivy": { "sbomqs score": 7, ... }, ... }
    """
    col_to_tool = {}
    current_tool = None
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            current_tool = str(value).strip().lower().replace(" ", "")
        if current_tool and col > 1:
            col_to_tool[col] = current_tool

    column_map = {}
    for col in range(2, ws.max_column + 1):
        sub_header = ws.cell(row=2, column=col).value
        tool = col_to_tool.get(col)
        if sub_header and tool:
            sub = str(sub_header).strip().lower()
            column_map.setdefault(tool, {})[sub] = col

    return column_map


def get_tool_name(filename):
    """
    Converts a filename to the tool name used in the Excel header.
      trivy.json             -> trivy
      syft.json              -> syft
      cargosbom.json         -> cargosbom
      agent-browser.cdx.json -> cyclonedx  (all *.cdx.json = cyclonedx)
    """
    name = filename.lower()
    if "cargosbom"        in name: return "cargosbom"
    if "cyclonedx"        in name: return "cyclonedx"
    if "trivy"            in name: return "trivy"
    if "syft"             in name: return "syft"
    if name.endswith(".cdx.json"):  return "cyclonedx"
    return name.split(".")[0]


def expand_path(relative_path):
    """
    Prepends BASE_PATH to a relative project path.
    """
    p = relative_path.strip()
    if p.startswith("/"):
        return p
    return f"{BASE_PATH}/{p}"


def parse_basic_scores(path):
    """
    Reads the output of: sbomqs score --basic --recursive <folder>
    Returns: { ("project_name", "tool_name"): score }
    """
    results = {}

    with open(path) as f:
        raw = f.read()

    for block in raw.split("---END---"):
        project = ""
        for line in block.splitlines():
            if line.startswith("PROJECT="):
                project = line.replace("PROJECT=", "").strip()
                continue

            if not line.strip() or "ERROR" in line or line.startswith("20"):
                continue

            cols = [c.strip() for c in line.split("\t")]

            if len(cols) < 6 or not re.match(r"^[\d.]+$", cols[0]):
                continue

            score    = float(cols[0])
            filepath = cols[5]
            filename = filepath.split("/")[-1]
            tool     = get_tool_name(filename)

            results[(project, tool)] = score

    return results


def parse_bsi_scores(path):
    """
    Reads the output of: sbomqs score --profile bsi-v2.1 <file>
    """
    results = {}

    with open(path) as f:
        raw = f.read()

    for block in raw.split("---END---"):
        project = ""
        tool    = ""
        body    = []

        for line in block.splitlines():
            if line.startswith("PROJECT="):
                project = line.replace("PROJECT=", "").strip()
            elif line.startswith("TOOL="):
                tool = line.replace("TOOL=", "").strip()
            else:
                body.append(line)

        if not project or not tool:
            continue

        text = "\n".join(body)

        score_match = re.search(r"SBOM Quality Score:\s*([\d.]+)/([\d.]+)\s+Grade:\s*(\S+)", text)
        req_match   = re.search(r"Required Fields\s*:\s*(\d+/\d+)\s+compliant",   text)
        add_match   = re.search(r"Additional Fields\s*:\s*(\d+/\d+)\s+compliant", text)
        opt_match   = re.search(r"Optional Fields\s*:\s*(\d+/\d+)\s+present",     text)

        if score_match:
            score = score_match.group(1)
            max_s = score_match.group(2)
            grade = score_match.group(3)

            results[(project, tool)] = {
                "score_and_grade": f"{score}/{max_s} ({grade})",
                "req": req_match.group(1) if req_match else "",
                "add": add_match.group(1) if add_match else "",
                "opt": opt_match.group(1) if opt_match else "",
            }

    return results


def write_to_excel(excel_path, basic_scores, bsi_scores):
    """
    Opens the existing Excel file, clears old data, and writes one row per project.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    column_map = read_column_map(ws)

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.value = None

    all_projects = list(dict.fromkeys(
        project for (project, _) in list(basic_scores.keys()) + list(bsi_scores.keys())
    ))

    for row_number, project in enumerate(all_projects, start=3):
        ws.cell(row=row_number, column=1, value=project)

        for tool, sub_columns in column_map.items():
            basic_val = basic_scores.get((project, tool), "")
            bsi_val   = bsi_scores.get((project, tool), {})

            values = {
                "sbomqs score": basic_val,
                "bsi-v2.1":     bsi_val.get("score_and_grade", ""),
                "required":     bsi_val.get("req", ""),
                "additional":   bsi_val.get("add", ""),
                "optional":     bsi_val.get("opt", ""),
            }

            for sub_name, col_number in sub_columns.items():
                ws.cell(row=row_number, column=col_number, value=values.get(sub_name, ""))

    tmp = excel_path + ".tmp"
    wb.save(tmp)
    shutil.move(tmp, excel_path)

    print(f"Done! Wrote {len(all_projects)} project(s) to {excel_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python3 parse_sbomqs.py <basic_tmp> <bsi_tmp> <excel_file>")
        sys.exit(1)

    basic_tmp, bsi_tmp, excel_file = sys.argv[1], sys.argv[2], sys.argv[3]

    print("Parsing basic scores ...")
    basic_scores = parse_basic_scores(basic_tmp)

    print("Parsing BSI v2.1 scores ...")
    bsi_scores = parse_bsi_scores(bsi_tmp)

    print("Writing to Excel ...")
    try:
        write_to_excel(excel_file, basic_scores, bsi_scores)
    except PermissionError:
        print(f"\n  ERROR: Cannot write to {excel_file}")
        print("  -> Close the file in Excel first, then run the script again.")
        sys.exit(1)
    except FileNotFoundError:
        print(f"\n  ERROR: Excel file not found: {excel_file}")
        print("  -> Make sure sbomqs_results.xlsx is in the same folder as this script.")
        sys.exit(1)