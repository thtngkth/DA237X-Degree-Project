"""
Compares Cargo metadata JSON with an SBOM, reporting on missing packages,
version mismatches, false dependencies, transitive dependency coverage,
and precision, recall, and F1 scores.
"""

import os
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from cargo import CargoLock
from sbom import SBOM


def safe_precision(tp, fp):
    return 0.0 if tp + fp == 0 else tp / (tp + fp)


def safe_recall(tp, fn):
    return 0.0 if tp + fn == 0 else tp / (tp + fn)


def safe_f1(precision, recall):
    return 0.0 if precision + recall == 0 else 2 * precision * recall / (precision + recall)


def fmt(value):
    """Format a percentage value with two decimals."""
    return f"{value:.2f}"


def pct(count, total):
    """Return percentage of count / total."""
    if total == 0:
        return "0.00"
    return fmt((count / total) * 100)


class Comparator:
    """Compare Cargo metadata JSON with SBOM."""
    def __init__(self, cargo: CargoLock, sbom: SBOM):
        self.cargo = cargo
        self.sbom = sbom

    def calculate_transitive(self, adj, start):
        """Walk the dependency graph from 'start' and return all reachable nodes."""
        visited = set()
        stack = [start]

        while stack:
            node = stack.pop()
            if node in visited:
                continue

            visited.add(node)
            stack.extend(adj.get(node, []))

        visited.discard(start)
        return visited

    def compare(self):
        """Run all comparisons and compute descriptive stats plus P/R/F1."""
        self.total_cargo_packages = len(self.cargo.packages)
        self.total_sbom_components = len(self.sbom.packages)

        cargo_components = self.cargo.component_set()
        sbom_components = self.sbom.component_set()

        cargo_names = {name for name, _ in cargo_components}    # set of (name, version)
        sbom_names = {name for name, _ in sbom_components}      # set of (name, version)

        lookup = self.cargo.build_lookup()                      # { name: {version: ...} }

        # Correct packages = Packages where both name AND version match exactly
        self.correctly_identified = cargo_components & sbom_components

        # Cargo version mismatches = (name, version) pairs exist in the SBOM but the exact version is wrong
        self.cargo_version_mismatched = {
            (name, version)
            for name, version in cargo_components
            if name in sbom_names and (name, version) not in sbom_components
        }

        # Missed components = Packages in Cargo doesn't appear in the SBOM
        self.missed_components = {
            (name, version)
            for name, version in cargo_components
            if name not in sbom_names
        }

        # Invented packages = Packages in the SBOM but not in Cargo
        seen_invented = set()
        self.missing_in_cargo = []

        for pkg in self.sbom.packages:
            if pkg.name not in cargo_names and pkg.name not in seen_invented:
                seen_invented.add(pkg.name)
                self.missing_in_cargo.append(pkg.id())

        # SBOM version mismatches = (name, version) pairs exist in Cargo but the exact version is wrong
        seen_mismatches = set()
        self.version_mismatches = []

        for pkg in self.sbom.packages:
            if pkg.name in lookup and pkg.version not in lookup[pkg.name]:
                key = pkg.id()
                if key not in seen_mismatches:
                    seen_mismatches.add(key)
                    self.version_mismatches.append(key)

        # Edge coverage
        cargo_edges = self.cargo.edges()
        sbom_edges = self.sbom.edges()

        correct_edges = cargo_edges & sbom_edges
        self.missing_edges = cargo_edges - sbom_edges   # in Cargo, absent from SBOM
        self.false_edges = sbom_edges - cargo_edges     # in SBOM,  absent from Cargo

        self.edge_coverage_pct = (
            len(correct_edges) / len(cargo_edges) * 100 if cargo_edges else 0.0
        )

        # Transitive dependency coverage
        cargo_adj = self.cargo.adjacency()
        sbom_adj = self.sbom.adjacency()

        self.transitive_missing_hist = defaultdict(list)    # list of packages with n missing transitive deps
        self.transitive_totals = {}                         # pkg -> total transitive dep count (Cargo)

        # Average per-package transitive coverage score
        coverage_scores = []

        for pkg in cargo_adj.keys():
            cargo_trans = self.calculate_transitive(cargo_adj, pkg)
            sbom_trans = self.calculate_transitive(sbom_adj, pkg)

            missing = cargo_trans - sbom_trans
            total = len(cargo_trans)

            self.transitive_missing_hist[len(missing)].append(pkg)
            self.transitive_totals[pkg] = total

            score = (total - len(missing)) / total * 100 if total > 0 else 100.0
            coverage_scores.append(score)

        self.avg_transitive_coverage = (
            sum(coverage_scores) / len(coverage_scores) if coverage_scores else 0.0
        )

        # Count packages with bucket transitive coverage
        self.trans_0_20 = sum(1 for s in coverage_scores if 0 <= s < 20)
        self.trans_20_40 = sum(1 for s in coverage_scores if 20 <= s < 40)
        self.trans_40_60 = sum(1 for s in coverage_scores if 40 <= s < 60)
        self.trans_60_80 = sum(1 for s in coverage_scores if 60 <= s < 80)
        self.trans_80_100 = sum(1 for s in coverage_scores if 80 <= s <= 100)

        """Component detection
        TP: name present in both Cargo and SBOM
        FP: name in SBOM but not in Cargo  (invented)
        FN: name in Cargo but not in SBOM  (missed)"""
        comp_tp = len(cargo_names & sbom_names)
        comp_fp = len(sbom_names - cargo_names)
        comp_fn = len(cargo_names - sbom_names)

        self.comp_precision = safe_precision(comp_tp, comp_fp)
        self.comp_recall = safe_recall(comp_tp, comp_fn)
        self.comp_f1 = safe_f1(self.comp_precision, self.comp_recall)

        """ Version accuracy - package name must exist in Cargo
        TP: name AND version match exactly
        FP: name exists in Cargo but SBOM reports a wrong version
        FN: correct Cargo version absent from SBOM"""
        ver_tp = 0
        ver_fp = 0
        seen_ver = set()

        for pkg in self.sbom.packages:
            if pkg.name in lookup:
                key = (pkg.name, pkg.version)

                if key in seen_ver:
                    continue

                seen_ver.add(key)

                if pkg.version in lookup[pkg.name]:
                    ver_tp += 1
                else:
                    ver_fp += 1

        ver_fn = len({
            (name, version)
            for name, version in cargo_components
            if name in sbom_names and (name, version) not in sbom_components
        })

        self.ver_precision = safe_precision(ver_tp, ver_fp)
        self.ver_recall = safe_recall(ver_tp, ver_fn)
        self.ver_f1 = safe_f1(self.ver_precision, self.ver_recall)

        """Direct dependency edges
        TP: edge present in both Cargo and SBOM
        FP: edge in SBOM but not in Cargo  (false edge)
        FN: edge in Cargo but not in SBOM  (missing edge)"""
        edge_tp = len(correct_edges)
        edge_fp = len(self.false_edges)
        edge_fn = len(self.missing_edges)

        self.edge_precision = safe_precision(edge_tp, edge_fp)
        self.edge_recall = safe_recall(edge_tp, edge_fn)
        self.edge_f1 = safe_f1(self.edge_precision, self.edge_recall)

        # Transitive dependency relationships. Include packages only in SBOM (contributes to FP transitive dep)
        trans_tp = 0
        trans_fp = 0
        trans_fn = 0

        all_roots = set(cargo_adj.keys()) | set(sbom_adj.keys())

        for pkg in all_roots:
            cargo_trans = self.calculate_transitive(cargo_adj, pkg)
            sbom_trans = self.calculate_transitive(sbom_adj, pkg)

            trans_tp += len(cargo_trans & sbom_trans)
            trans_fp += len(sbom_trans - cargo_trans)
            trans_fn += len(cargo_trans - sbom_trans)

        self.trans_precision = safe_precision(trans_tp, trans_fp)
        self.trans_recall = safe_recall(trans_tp, trans_fn)
        self.trans_f1 = safe_f1(self.trans_precision, self.trans_recall)

    @staticmethod
    def clear_report(path):
        """Create or reset the Excel file and write the header row."""
        headers = [
            "Project",
            "Cargo pkg",
            "SBOM pkg",
            
            "Correct pkg",
            "Cargo version mismatch",
            "Edg coverage (%)",
            "Edg missing (%)",
            "False edg (%)",
            "Missed comp",
            "Invented pkg",
            "SBOM version mismatch",
            "Avg trans dep (%)",
            "0-20% trans dep",
            "20-40% trans dep",
            "40-60% trans dep",
            "60-80% trans dep",
            "80-100% trans dep",

            "Comp Precision (%)",
            "Comp Recall (%)",
            "Comp F1 (%)",

            "Ver Precision (%)",
            "Ver Recall (%)",
            "Ver F1 (%)",

            "Edge Precision (%)",
            "Edge Recall (%)",
            "Edge F1 (%)",
            
            "Trans Precision (%)",
            "Trans Recall (%)",
            "Trans F1 (%)",
        ]

        wb = Workbook()
        sheet = wb.active
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(name="Arial", bold=True)

        wb.save(path)
        print(f"Created/reset report with headers: {path}")

    def write_report(self, path, project_name):
        if not os.path.exists(path):
            Comparator.clear_report(path)

        wb = load_workbook(path)
        sheet = wb.active

        cargo_edges_total = len(self.cargo.edges())
        sbom_edges_total = len(self.sbom.edges())
        cargo_components_total = len(self.cargo.component_set())
        total_pkgs = len(self.transitive_totals)

        new_row = [
            project_name,
            self.total_cargo_packages,
            self.total_sbom_components,

            len(self.correctly_identified),
            len(self.cargo_version_mismatched),
            fmt(self.edge_coverage_pct),
            pct(len(self.missing_edges), cargo_edges_total),
            pct(len(self.false_edges), sbom_edges_total),
            pct(len(self.missed_components), cargo_components_total),
            pct(len(self.missing_in_cargo), self.total_sbom_components),
            pct(len(self.version_mismatches), self.total_sbom_components),
            fmt(self.avg_transitive_coverage),
            pct(self.trans_0_20, total_pkgs),
            pct(self.trans_20_40, total_pkgs),
            pct(self.trans_40_60, total_pkgs),
            pct(self.trans_60_80, total_pkgs),
            pct(self.trans_80_100, total_pkgs),

            fmt(self.comp_precision * 100),
            fmt(self.comp_recall * 100),
            fmt(self.comp_f1 * 100),

            fmt(self.ver_precision * 100),
            fmt(self.ver_recall * 100),
            fmt(self.ver_f1 * 100),

            fmt(self.edge_precision * 100),
            fmt(self.edge_recall * 100),
            fmt(self.edge_f1 * 100),

            fmt(self.trans_precision * 100),
            fmt(self.trans_recall * 100),
            fmt(self.trans_f1 * 100),
        ]

        row_index = sheet.max_row + 1

        for col_index, value in enumerate(new_row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

        for col in range(1, len(new_row) + 1):
            sheet.cell(row=row_index, column=col).font = Font(name="Arial")

        wb.save(path)